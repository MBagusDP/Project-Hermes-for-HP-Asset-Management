import os
import pandas as pd
import openpyxl
from skimpy import clean_columns
import gspread
import sys
import time
from datetime import datetime
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException, ElementNotInteractableException, TimeoutException
import warnings
warnings.filterwarnings("ignore")

# Globally define 'try_count'
try_count = 0

if __name__ == "__main__":

    print("\n  ##### READ ME! #####")
    time.sleep(3)
    print("   This is an automated browser task for downloading fanpage karma metrics data")
    time.sleep(5)
    print("   As google security system is pretty tight")
    time.sleep(5)
    print("   This task will have to be executed with browser UI displayed on the screen")
    time.sleep(5)
    print("   Duration of the program depends on the number of data that the program have to download")
    time.sleep(5)
    print("   Please put your hands off from your keyboard and mouse for as long as the program is running")
    time.sleep(10)
    print("\n  Thank you for your attention...")
    time.sleep(10)

    print("\n##### Initiating automated browser tasks for fanpage karma #####")

    # Obtain the latest date from the data
    print("\nFinding existing CSV file...")
    if os.path.exists("D:\\HPAM\\HPAM_projects\\Hermes\\fanpage_karma_CSV\\Explore_Mohammad_Bagus_Dwi_Putra.csv") == False:
        print("   Existing CSV file not found! setting up date to the earliest downloadable date in fanpage karma...")
        latest_date = pd.to_datetime("2018-05-16").strftime("%m/%d/%y")
        print("   Date has been set to 2018-05-16!")
    else:
        print("   Existing CSV file found! setting up date to the latest date found in the data...")
        date_df = clean_columns(pd.read_csv("fanpage_karma_CSV/Explore_Mohammad_Bagus_Dwi_Putra.csv"))
        latest_date = datetime.strptime(date_df.date.max(), "%Y-%m-%d").strftime("%m/%d/%y")
        print("   Date has been set to", str(datetime.strptime(date_df.date.max(), "%Y-%m-%d").strftime("%Y-%m-%d")) + "!")

    # When the latest date in existing data is the date of today, exit python program
    if latest_date != "05/16/18":
        print("\nChecking if the data is up to date...")
        if latest_date == datetime.date(datetime.now()).strftime("%m/%d/%y"):
            print("   Data is up to date!")
            print("\n##### Task completed! #####")
            sys.exit()
        print("   Possible update found!")
        print("   Update to be made: " + str(datetime.strptime(date_df.date.max(), "%Y-%m-%d").strftime("%Y-%m-%d")) +\
            " to " + str(datetime.date(datetime.now()).strftime("%Y-%m-%d")))

    # Remove Excel file from downloads folder if exist
    print("\nFinding previously downloaded Excel file on the Downloads folder...")
    file_count = 0
    for file in os.listdir(r"C:\Users\HPAM\Downloads"):
        if "Explore_Mohammad_Bagus_Dwi_Putra" in file:
            file_count = file_count + 1
            os.remove("C:\\Users\\HPAM\\Downloads\\" + file)
    if file_count > 0:
        if file_count == 1:
            # Singular form of grammar
            print("   " + str(file_count) + " Excel file found!")
            print("   Excel file has been removed!")
        else:
            # Plural form of grammar
            print("   " + str(file_count) + " Excel files found!")
            print("   Excel files has been removed!")
    else:
        print("   No previously downloaded Excel file found!")

    # Initialize chrome
    print("\nStarting browser tasks...")
    print("   Initializing Chrome browser...")
    driver = uc.Chrome(version_main=99)
    
    # Define functions
    def wait_for_visibility(tag_name, elem_name, try_count):
        switcher = {
            "ID": By.ID,
            "NAME": By.NAME,
            "CLASS_NAME": By.CLASS_NAME,
            "XPATH": By.XPATH,
            "CSS_SELECTOR": By.CSS_SELECTOR,
            "TAG_NAME": By.TAG_NAME
        }
        try:
            WebDriverWait(driver, 10).until(EC.visibility_of_element_located((switcher.get(tag_name), elem_name)))
        except TimeoutException:
            if try_count > 9:
                print("\nPage reloaded more than 10 times!")
                print("\nProgram terminated!")
                sys.exit()
            print("Error: Element '" + elem_name + "' not found, reloading page...")
            return "Timeout"

    def wait_for_invisibility(tag_name, elem_name, try_count):
        switcher = {
            "ID": By.ID,
            "NAME": By.NAME,
            "CLASS_NAME": By.CLASS_NAME,
            "XPATH": By.XPATH,
            "CSS_SELECTOR": By.CSS_SELECTOR,
            "TAG_NAME": By.TAG_NAME
        }
        try:
            WebDriverWait(driver, 10).until(EC.invisibility_of_element_located((switcher.get(tag_name), elem_name)))
        except TimeoutException:
            if try_count > 9:
                print("\nPage reloaded more than 10 times!")
                print("\nProgram terminated!")
                sys.exit()
            print("Error: Element '" + elem_name + "' still present after 10 seconds, reloading page...")
            return "Timeout"

    def click_elem(tag_name, elem_name, try_count):
        switcher = {
            "ID": By.ID,
            "NAME": By.NAME,
            "CLASS_NAME": By.CLASS_NAME,
            "XPATH": By.XPATH,
            "CSS_SELECTOR": By.CSS_SELECTOR,
            "TAG_NAME": By.TAG_NAME
        }
        try:
            element = WebDriverWait(driver, 10).\
                until(EC.element_to_be_clickable((switcher.get(tag_name), elem_name)))
            try:
                element.click()
            except StaleElementReferenceException:
                if try_count > 9:
                    print("\nPage reloaded more than 10 times!")
                    print("\nProgram terminated!")
                    sys.exit()
                print("Error: Element '" + elem_name + "' not attached to the page document, reloading page...")
                return "Timeout"
            except ElementNotInteractableException:
                if try_count > 9:
                    print("\nPage reloaded more than 10 times!")
                    print("\nProgram terminated!")
                    sys.exit()
                print("Error: Element '" + elem_name + "' not interactable, reloading page...")
                return "Timeout"
        except TimeoutException:
            if try_count > 9:
                print("\nPage reloaded more than 10 times!")
                print("\nProgram terminated!")
                sys.exit()
            print("Error: Element '" + elem_name + "' not clickable, reloading page...")
            return "Timeout"

    # Get ID and PASSWORD
    #   As ID and PASSWORD is a private data, I do not explicitly embed them inside this code...
    #   Please follow this steps to enable you to login to google account with your own account:
    #       1. Create your own .txt file contaning your ID and PASSWORD separated with a coma
    #               example: "someone@gmail.com,YourPassword123"
    #       2. Put your .txt inside a folder named 'google_account'
    #       3. Change 'idpass_hpfinancials.txt' into your own named .txt file
    #       4. Steps complete, cheers!
    identity = open("google_account/idpass_hpfinancials.txt", "r").read().split(",")[0]
    passcode = open("google_account/idpass_hpfinancials.txt", "r").read().split(",")[1]

    # login to google account
    print("   Login to Google account...")
    driver.get("https://accounts.google.com")
    while True:
        if wait_for_visibility("ID", "identifierId", try_count) != "Timeout":
            driver.find_element(By.ID, "identifierId").send_keys(identity)
            click_elem("CLASS_NAME", "VfPpkd-vQzf8d", try_count)
            try_count = 0
        else:
            try_count = try_count + 1
            driver.get("https://accounts.google.com")
            
        if wait_for_visibility("NAME", "password", try_count) != "Timeout":
            driver.find_element(By.NAME, "password").send_keys(passcode)
            click_elem("ID", "passwordNext", try_count)
            try_count = 0
            break
        else:
            try_count = try_count + 1
            driver.get("https://accounts.google.com")

    # Login to fanpage karma
    wait_for_visibility("CLASS_NAME", "x7WrMb", try_count)
    driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
    while True:
        if wait_for_visibility("ID", "googleLoginButton", try_count) != "Timeout":
            print("   Login to fanpage karma...")
            click_elem("ID", "googleLoginButton", try_count)
            try_count = 0
        else:
            if wait_for_visibility("XPATH", "//div[@class='col-6 col-lg-5 col-xl-4 navbar-sidebar-header']", try_count) != "Timeout":
                break
            try_count = try_count + 1
            driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
            continue

        # Navigate to fanpage karma post
        if wait_for_visibility("XPATH", "//div[@class='col-6 col-lg-5 col-xl-4 navbar-sidebar-header']", try_count) == "Timeout":
            try_count = try_count + 1
            driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
            continue

        break

    print("   Getting range of dates...")
    print("      Date range:", pd.to_datetime(latest_date).strftime("%Y-%m-%d"), "to", datetime.date(datetime.now()))
    print("\n   Downloading a total of", str(len(pd.date_range(latest_date, datetime.now()))) + " Excel files...")
    iter_track = 0
    for date in pd.date_range(latest_date, datetime.now()).strftime("%m/%d/%y"):
        iter_track = iter_track + 1
        while True:
            # Set datetime
            print("\n----------------------------   Iteration " + str(iter_track) + "   ------------------------------")
            print("      Filling up date range...")

            # Double check if previously downloaded file has been removed
            print("\n      Removing downloaded Excel file...")
            file_count = 0
            for file in os.listdir(r"C:\Users\HPAM\Downloads"):
                if "Explore_Mohammad_Bagus_Dwi_Putra" in file:
                    file_count = file_count + 1
                    os.remove("C:\\Users\\HPAM\\Downloads\\" + file)
                elif "Explore Mohammad Bagus Dwi Putra_" in file:
                    file_count = file_count + 1
                    os.remove("C:\\Users\\HPAM\\Downloads\\" + file)
            if file_count > 0:
                print("         File has been removed!")

            wait_for_visibility("CLASS_NAME", "navbar-sidebar-select-text", try_count)
            while True:
                try:
                    driver.find_element(By.CLASS_NAME, "navbar-sidebar-select-text").text != "EXPLORE MOHAMMAD BAGUS DWI PUTRA (5)"
                    break
                except:
                    "loop until false"

            if wait_for_visibility("CLASS_NAME", "js-benchmarkingPanel-download-link", try_count) != "Timeout":
                "continue to the next task"
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            if wait_for_visibility("CLASS_NAME", "datepickerZeitraumText", try_count) != "Timeout":
                click_elem("CLASS_NAME", "datepickerZeitraumText", try_count)
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            print("         Filling date with :", pd.to_datetime(date).strftime("%Y-%m-%d") + "...")
            if wait_for_visibility("XPATH", "//input[@class='form-control daterangepicker-field daterangepicker-field--base js-baseDatesVonInput']", try_count) != "Timeout":
                while True:
                    try:
                        driver.find_element(By.XPATH, "//input[@class='form-control daterangepicker-field daterangepicker-field--base js-baseDatesVonInput']").\
                            clear()
                        driver.find_element(By.XPATH, "//input[@class='form-control daterangepicker-field daterangepicker-field--base js-baseDatesVonInput']").\
                            send_keys(date)
                        break
                    except ElementNotInteractableException:
                        "try again"
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            if wait_for_visibility("XPATH", "//input[@class='form-control daterangepicker-field daterangepicker-field--base js-baseDatesBisInput']", try_count) != "Timeout":
                while True:
                    try:
                        driver.find_element(By.XPATH, "//input[@class='form-control daterangepicker-field daterangepicker-field--base js-baseDatesBisInput']").\
                            clear()
                        driver.find_element(By.XPATH, "//input[@class='form-control daterangepicker-field daterangepicker-field--base js-baseDatesBisInput']").\
                            send_keys(date)
                        break
                    except ElementNotInteractableException:
                        "try again"
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue
            
            if wait_for_visibility("CSS_SELECTOR", ".popupFooter.text-center", try_count) != "Timeout":
                parent = driver.find_element(By.CSS_SELECTOR, ".popupFooter.text-center")
                WebDriverWait(parent, 10).until(EC.visibility_of_element_located((By.TAG_NAME, "button")))
                elem = parent.find_elements(By.TAG_NAME, "button")
                click_elem("ID", elem[0].get_attribute("id"), try_count)
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")

            # Download Excel file
            print("      Downloading Excel file...")
            while True:
                try:   
                    while driver.find_element(By.CSS_SELECTOR, ".infoPopupBorder__content.infoPopupBorder__content--fixed").is_displayed() == True:
                        "loop until display is closed"
                    break
                except StaleElementReferenceException:
                    "loop until false"

            if wait_for_visibility("CLASS_NAME", "js-benchmarkingPanel-download-link", try_count) != "Timeout":
                click_elem("CLASS_NAME", "js-benchmarkingPanel-download-link", try_count)
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            if wait_for_visibility("CSS_SELECTOR", ".dropdown-menu.dropdown-menu-right.darkDropDown.show", try_count) != "Timeout":
                parent = driver.find_element(By.CSS_SELECTOR, ".dropdown-menu.dropdown-menu-right.darkDropDown.show")
                wait_for_visibility("TAG_NAME", "a", try_count)
                elem = parent.find_elements(By.TAG_NAME, "a")
                click_elem("ID", elem[1].get_attribute("id"), try_count)
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue
            
            # Check if download is successful
            if wait_for_visibility("CLASS_NAME", "downloadOverlay--show", try_count) != "Timeout":
                "go to next task"
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            if wait_for_invisibility("CLASS_NAME", "downloadOverlay--show", try_count) != "Timeout":
                "go to next task"
            else:
                try_count = try_count + 1
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            try_read = 0
            while True:
                loop = "continue"
                for file in os.listdir(r"C:\Users\HPAM\Downloads"):
                    if "Explore Mohammad Bagus Dwi Putra_" in file:
                        try:
                            openpyxl.load_workbook("C:\\Users\\HPAM\\Downloads\\" + file)
                            print("         File has been successfully downloaded!")
                            loop = "break"
                            break
                        except:
                            time.sleep(0.1)
                            try_read = try_read + 1
                            if try_read > 100:
                                loop = "break"
                            break
                if loop == "break":
                    break
            if try_read > 200:
                print("Error: downloaded file is corrupted! reloading page...")
                driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                continue

            # Renaming the file
            print("\n      Starting data processing...")
            print("         Renaming downloaded Excel file...")
            if os.path.exists("D:\\HPAM\\HPAM_projects\\Hermes\\fanpage_karma_CSV\\Explore_Mohammad_Bagus_Dwi_Putra.csv") == False:
                for file in os.listdir(r"C:\Users\HPAM\Downloads"):
                    if "Explore Mohammad Bagus Dwi Putra_" in file:
                        os.rename("C:\\Users\\HPAM\\Downloads\\" + file,\
                                "C:\\Users\\HPAM\\Downloads\\Explore_Mohammad_Bagus_Dwi_Putra.xlsx")
                        break
            else:
                for file in os.listdir(r"C:\Users\HPAM\Downloads"):
                    if "Explore Mohammad Bagus Dwi Putra_" in file:
                        os.rename("C:\\Users\\HPAM\\Downloads\\" + file,\
                                "C:\\Users\\HPAM\\Downloads\\Explore_Mohammad_Bagus_Dwi_Putra_"\
                                + str(datetime.date(datetime.now())) + ".xlsx")
                        break

            # Processing the CSV file
            print("         Importing downloaded Excel file...")
            if os.path.exists("D:\\HPAM\\HPAM_projects\\Hermes\\fanpage_karma_CSV\\Explore_Mohammad_Bagus_Dwi_Putra.csv") == False:
                try:
                    xl_workbook = openpyxl.load_workbook("C:\\Users\\HPAM\\Downloads\\Explore_Mohammad_Bagus_Dwi_Putra.xlsx")
                except:
                    print("Error: renamed file is corrupted! reloading page...")
                    driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                    continue
                sheets = xl_workbook.sheetnames
                xl_sheet = xl_workbook[sheets[0]]
            else:
                try:
                    xl_workbook = openpyxl.load_workbook("C:\\Users\\HPAM\\Downloads\\Explore_Mohammad_Bagus_Dwi_Putra_"\
                                                + str(datetime.date(datetime.now())) + ".xlsx")
                except:
                    print("Error: renamed file is corrupted! reloading page...")
                    driver.get("https://www.fanpagekarma.com/dashboard?h=cajB0qnXu")
                    continue
                sheets = xl_workbook.sheetnames
                xl_sheet = xl_workbook[sheets[0]]

            
            print("         Processing the imported CSV data...")
            xl_sheet.delete_cols(0)
            xl_sheet.delete_rows(0, 4)
            print("            Header removed!")
            df = pd.DataFrame(xl_sheet.values)
            xl_workbook.close()
            df.columns = df.iloc[0]
            df = clean_columns(df.drop(0).reset_index(drop=True))
            print("            Column names cleaned!")
            for i in range(0, len(df.columns)-8):
                if i == 0:
                    df = df.drop(columns=["header"])
                else:
                    df = df.drop(columns=["header_" + str(i)])
            df = df.dropna(how="all")
            print("            Rows and columns with empty cells removed!")
            df = df.fillna(0)
            df = df.replace("", 0).replace("-", 0)
            print("            Empty values has been filled with zeroes!")
            df["date"] = pd.to_datetime(date).strftime("%Y-%m-%d")
            print("            Column 'date' has been filled with", pd.to_datetime(date).strftime("%Y-%m-%d"))
            print("      Data successfully processed!")

            # Check whether the data exist in the existing data
            print("\n      Checking the existance of the downloaded data inside the existing data table...")
            if os.path.exists("D:\\HPAM\\HPAM_projects\\Hermes\\fanpage_karma_CSV\\Explore_Mohammad_Bagus_Dwi_Putra.csv") == True:
                existing_df = pd.read_csv("D:\\HPAM\\HPAM_projects\\Hermes\\fanpage_karma_CSV\\Explore_Mohammad_Bagus_Dwi_Putra.csv")
                if date == datetime.strptime(existing_df.date.max(), "%Y-%m-%d").strftime("%m/%d/%y"):
                    print("         Data found, no update to be made!")
                    # Remove file in downloads folder after we successfully modify and move the file
                    print("\n      Removing downloaded CSV file...")
                    file_count = 0
                    for file in os.listdir(r"C:\Users\HPAM\Downloads"):
                        if "Explore_Mohammad_Bagus_Dwi_Putra" in file:
                            file_count = file_count + 1
                            os.remove("C:\\Users\\HPAM\\Downloads\\" + file)
                    if file_count > 0:
                        print("         File has been removed!")
                    break
            print("         Data not found!")

            # Saving/updating the file as CSV
            if os.path.exists("D:\\HPAM\\HPAM_projects\\Hermes\\fanpage_karma_CSV\\Explore_Mohammad_Bagus_Dwi_Putra.csv") == False:
                print("\n      Saving data as CSV file...")
                df.to_csv("fanpage_karma_CSV/Explore_Mohammad_Bagus_Dwi_Putra.csv", index=False)
                print("         CSV file has been successfully created!")
            else:
                print("\n      Updating existing CSV file...")
                new_df = pd.concat([df, existing_df]).reset_index(drop=True)
                new_df.to_csv("fanpage_karma_CSV/Explore_Mohammad_Bagus_Dwi_Putra.csv", index=False)
                print("         Existing CSV file has been successfully updated!")

            # Remove file in downloads folder after we successfully modify and move the file
            print("\n      Removing downloaded Excel file...")
            file_count = 0
            for file in os.listdir(r"C:\Users\HPAM\Downloads"):
                if "Explore_Mohammad_Bagus_Dwi_Putra" in file:
                    file_count = file_count + 1
                    os.remove("C:\\Users\\HPAM\\Downloads\\" + file)
            if file_count > 0:
                print("         File has been removed!")

            # Break from while loop if task run sucessfully
            break

    # Ending browser tasks
    print("\nBrowser task has been sucessfully executed!")
    print("Ending browser task...")
    driver.quit()

    # Recalculating follower growth
    print("\nRecalculating follower growth...")
    df = pd.read_csv("fanpage_karma_CSV/Explore_Mohammad_Bagus_Dwi_Putra.csv")
    HPAM_IG = df[(df.page == "HP Asset Management") & (df.network == "INSTAGRAM")].reset_index(drop=True)
    HPAM_FB = df[(df.page == "HP Asset Management") & (df.network == "FACEBOOK")].reset_index(drop=True)
    HPF = df[df.page == "HP Financials"].reset_index(drop=True)
    MyHero = df[df.page == "MyHero by HPAM"].reset_index(drop=True)

    HPAM_IG.follower_growth_absolute = HPAM_IG.sort_values(by="date").fans.diff()
    HPAM_FB.follower_growth_absolute = HPAM_FB.sort_values(by="date").fans.diff()
    HPF.follower_growth_absolute = HPF.sort_values(by="date").fans.diff()
    MyHero.follower_growth_absolute = MyHero.sort_values(by="date").fans.diff()
    df = pd.concat([HPAM_IG, HPAM_FB, HPF, MyHero]).sort_values(by="date", ascending=False).fillna(0)
    df.to_csv("fanpage_karma_CSV/Explore_Mohammad_Bagus_Dwi_Putra.csv", index=False)
    print("   Follower growth has been sucessfully recalculated!")

    # Saving/updating the file as csv and send it to google sheet
    print("\nUploading data to Google Sheet...")
    # create your own credetial key
    # sh = gspread.service_account(filename="credential_key_google_api/hermes_credential.json").open("HPAM_hermes_metrics")
    sh.sheet1.update([df.columns.values.tolist()] + df.values.tolist())
    print("   Google Sheet has been successfully updated!")

    # Annonunce that the program run successfully
    print("\n##### Task completed! #####")